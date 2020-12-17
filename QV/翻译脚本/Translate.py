 # -*- coding: UTF-8 -*-
#hao.wang@qualvision.cn

ini='''Translate.ini Example
------------------------------------------------------
[input]
vcxproj=../../qt/VMSApplication/VMSApplication.vcxproj
[setting]
langCount=6
sysCmdPath=D:/Qt/5.13.0/msvc2017/bin/
deleteTemp=false
------------------------------------------------------'''

import os
import re
import sys
import shutil
import openpyxl
import configparser
from xml.etree.ElementTree import ElementTree,Element

titleList=[]

#excel文件和语言个数
#行和列的索引都是从1开始
def Excel2TrCpp(excel,langCount):
    '''将excel文件转换为对应语言版本的tr文件与Transkey.cpp
    excel:excel文件
    langCount：语言个数
    transKey:transKey输出目录
    '''
    wb = openpyxl.load_workbook(excel)
    sheet = wb['VMS-Pro']
    rcount = sheet.max_row

    #excel的第一行是抬头，第一列是key，从第二列开始是语言列表
    header = 1
    cppkey = False
    for col in range(2, int(langCount)+2):
        cell = sheet.cell(header,col).value
        if 0 == len(cell):
            continue

        title = cell.split("|")

        #储存qm以及中间文件的的默认名字
        #' '空格会导致cmd参数错误
        titleList.append(title[0].replace(' ','_'))

        file = title[0].replace(' ','_')+".tr"
        lang = title[0] if len(title) == 1 else title[1]

        with open(tempTrPath+file,"w",encoding = "utf-8") as f:
            #lines = ["%s=%s"%(row[0],row[col-1])  for row in sheet[2:rcount]]
            #print(lines)
            f.write("[info]\n")
            f.write("Language=%s\n"%lang)
            f.write("[String]\n\n")

            for row in range(2,rcount+1):
                if(sheet.cell(row,1).value is None):
                    continue
					
                line = "%s=%s\n"%(sheet.cell(row,1).value,sheet.cell(row,col).value)
                f.write(line)
            f.close()
            print(file+' generated')
        
        if not cppkey:
            cppkey = True
            with open(outputPath+'TransKey.cpp',"w",encoding = "utf-8") as f: 
                f.write('/*This temporary file is generated by Tanslate.py*/\n')
                f.write('#include "trtext.h"\n')
                f.write('void INFRA::TRTEXT::Refresh(){\n')
                for row in range(2,rcount+1):
                    if(sheet.cell(row,1).value is None):
                        continue
                        
                    line = 'm_str["%s"]=tr("%s");\n'%(sheet.cell(row,1).value,(sheet.cell(row,1).value))
                    f.write(line)
                f.write('}\n')
                f.close
                print('TransKey.cpp generated')


def Vs2Qt(vcxproj):
    '''提取vcxproj文件中的文件（.h .cpp .ui .qrc）至qt pro中
    qt的lupdate与lrelease命令就是对这些内容进行检索更新
    vcxproj：vs项目文件
    '''
    with open(tempQtPath + "temp.pro","w",encoding = "utf-8") as f:
        f.write('#This temporary file is generated by Translate.py\n\n')

        headerslist = []
        headers = ''
        sourceslist = []
        sources = ''
        formslist = []
        forms = ''
        reslist = []
        res = ''

        #对vcxproj中每一行数据用正则表带是提取并保存
        for line in open(vcxproj, encoding="utf-8"):
            headerslist += re.findall(r'Include="(.*?).h" />', line)
            sourceslist += re.findall(r'Include="(.*?).cpp" />', line)
            formslist += re.findall(r'Include="(.*?).ui" />', line)
            reslist += re.findall(r'Include="(.*?).qrc" />', line)

        #transTemp导致temp.pro文件相比于vcxproj文件目录加深一层，所以加上../回退至与vcxproj相同目录,以保证文件路径正确
        for item in headerslist:
            headers += ('../'+str(item)).replace('\\', '/')
            headers += '.h \\\n    '
        f.write('HEADERS += ')
        f.write(headers[:-4])

        for item in sourceslist:
            sources += ('../'+str(item)).replace('\\', '/')
            sources += '.cpp \\\n    '
        f.write('SOURCES += ')
        f.write(sources[:-4])

        for item in formslist:
            forms += ('../'+str(item)).replace('\\', '/')
            forms += '.ui \\\n    '
        f.write('FORMS += ')
        f.write(forms[:-4])

        for item in reslist:
            res += ('../'+str(item)).replace('\\', '/')
            res += '.qrc \\\n    '
        f.write('RESOURCES += ')
        f.write(res[:-4])

        f.close()
        print('temp.pro generated')


def read_xml(in_path):
    '''读取并解析xml文件
    in_path: xml路径
    return: ElementTree
    '''
    tree = ElementTree()
    tree.parse(in_path)
    return tree
 
def write_xml(tree, out_path):
    '''将xml文件写出
    tree: xml树
    out_path: 写出路径
    '''
    tree.write(out_path, encoding="utf-8",xml_declaration=True)

def find_nodes(tree, path):
    '''查找某个路径匹配的所有节点
    tree: xml树
    path: 节点路径
    '''
    return tree.findall(path)

def change_node_text(nodelist,tr_list):
    '''如果nodelist的source和tr_list的key匹配，将对应的value赋给translation
    nodelist:节点列表
    tr_list : source与translation（key,value）集合'''
    for tr_item in tr_list:
        for node in nodelist:
            if(tr_item[0] == node[len(node)-2].text):
                node[len(node)-1].text = tr_item[1]
                # break 有重复项，断开会导致填补不全

def read_tr(path):
    '''读取tr文件中的有效内容至列表中
    path:tr文件
    '''
    tr_file = open(path, mode="r", encoding="utf-8")
    tr_content = tr_file.readlines()
    # 定义空列表存储临时数据，将同一组数据存储在同一列表中
    tr_list = []
    for each in tr_content:
        each = each.replace("\n","")
        list = each.split("=")
        #成对出现 =
        if(len(list)==2):
            tr_list.append(list)

    tr_file.close()
    return tr_list

def MergeTrTs(tr,ts):
    '''将tr数据合并至ts文件中
    tr:tr文件
    ts:ts文件
    '''
    #读取tr文件
    tr_list = read_tr(tr)
    #读取ts文件
    ts_tree = read_xml(ts)
    #找到ts中应当存储tr数据的节点
    text_nodes = find_nodes(ts_tree, "context/message")
    #将tr数据存储至对应ts节点中
    change_node_text(text_nodes,tr_list)
    #覆盖保存原来的ts文件
    write_xml(ts_tree, ts)
    


if __name__ == "__main__":

    print('-'*35+'\n'+'\t'+'Program running\n'+'-'*35)

    #读取配置文件----
    curPath = os.path.dirname(os.path.realpath(__file__))
    configPath = os.path.join(curPath,"Translate.ini")
    if not os.path.exists(configPath):
        print('Error! Translate.ini not found!')
        print(ini)
        os.system('pause')
        os._exit()
    config = configparser.ConfigParser()
    print('Reading Translate.ini...')
    config.read(configPath,encoding="utf-8")

    #创建文件夹----
    tempTrPath ='./transTemp/'
    if not os.path.exists(tempTrPath):
        print('Create folder for temporary tr data :',tempTrPath)
        os.mkdir(tempTrPath)
    outputPath ='./transOutput/'
    if not os.path.exists(outputPath):
        print('Create folder for output data :',outputPath)
        os.mkdir(outputPath)

    #将Excel文件转换为临时Tr文件与Transkey.cpp文件----
    excel = config["input"]["excel"]
    print('Excel:',excel)
    langCount = config["setting"]["langCount"]
    print('Types of translation:',langCount)
    print('Excel 2 Tr CPP...')
    Excel2TrCpp(excel,langCount)

    #通过VS vcxproj建立临时QT的pro项目----
    vcxproj = config["input"]["vcxproj"]
    print('VS vcxproj: ',vcxproj)
    vcxprojDirName = os.path.dirname(vcxproj)+'/'
    tempQtPath = vcxprojDirName+'./transTemp/'
    if not os.path.exists(tempQtPath):
        print('Create folder for temporary Qt data :',tempQtPath)
        os.mkdir(tempQtPath)
    print('Vs 2 Qt...')
    Vs2Qt(vcxproj)


    #调用QT ludate命令生成临时TS文件----
    sysCmdPath = config['setting']['sysCmdPath']
    print('lupdate & lrelease:',sysCmdPath)

    for title in titleList:
        ts = tempQtPath + title + '.ts'
        ludateCmd = sysCmdPath+'lupdate '+tempQtPath+'temp.pro -ts '+ts
        print('ludate Cmd:',ludateCmd)
        os.system(ludateCmd)

        #将TR数据导入TS文件----
        tr =tempTrPath + title + '.tr'
        qm =outputPath + title + '.qm'
        MergeTrTs(tr,ts)
        print(title+'.tr '+title+'.ts merged')

        #调用QT lrelease命令生成QM文件----
        lreleaseCmd = sysCmdPath+'lrelease '+ts+' -qm '+qm
        print('lrelease Cmd:',lreleaseCmd)
        os.system(lreleaseCmd)
        print(title+'.qm generated')

    #是否删除中间临时文件----
    deleteTemp = config['setting']['deleteTemp']
    if deleteTemp=='true':
        shutil.rmtree(tempQtPath)
        shutil.rmtree(tempTrPath)
        print('All temporary data deleted')

    print('-'*35+'\n'+'\t'+'Program finished\n'+'-'*35)
    os.system('pause')